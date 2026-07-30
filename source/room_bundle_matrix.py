"""Excel matrix import/export for room-type asset-bundle assignments."""

from __future__ import annotations

from collections import defaultdict
from pathlib import Path
import re

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from xlsx_workbook import XlsxError, read_xlsx


MATRIX_SHEET_NAME = "Room Bundle Matrix"
ROWS_SHEET_NAME = "Assignment Rows"
REFERENCE_SHEET_NAME = "Reference"

_NAVY = "17365D"
_TEAL = "0F6B78"
_BLUE = "24557A"
_PALE_BLUE = "EAF2F8"
_PALE_YELLOW = "FFF3CD"
_WHITE = "FFFFFF"
_BORDER = Side(style="thin", color="D9E2F3")


def _text(value) -> str:
    return str(value if value is not None else "").strip()


def _key(value) -> str:
    return re.sub(r"[^a-z0-9]+", " ", _text(value).casefold()).strip()


def _room_id(room_type, index=0) -> str:
    return _text(room_type.get("id")) or f"ROOM-{index + 1}"


def _room_name(room_type) -> str:
    return _text(room_type.get("name")) or _text(room_type.get("id"))


def _bundle_id(bundle, index=0) -> str:
    return _text(bundle.get("id")) or f"BUNDLE-{index + 1}"


def _bundle_name(bundle) -> str:
    return _text(bundle.get("name")) or _text(bundle.get("id"))


def _quantity(value) -> int | None:
    if value in (None, ""):
        return 0
    if isinstance(value, bool):
        return 1 if value else 0
    if isinstance(value, (int, float)):
        number = int(value)
        return number if number >= 0 and float(value) == number else None
    text = _text(value)
    if not text:
        return 0
    if text.casefold() in {"x", "yes", "y", "true", "checked", "✓", "✔"}:
        return 1
    try:
        number = int(float(text))
    except ValueError:
        return None
    try:
        exact = float(text) == number
    except ValueError:
        exact = False
    return number if number >= 0 and exact else None


def _quantity_lookup(quantities, room_id, bundle_id) -> int:
    if isinstance(quantities, dict):
        if (room_id, bundle_id) in quantities:
            return max(0, int(quantities[(room_id, bundle_id)] or 0))
        room_values = quantities.get(room_id, {})
        if isinstance(room_values, dict):
            return max(0, int(room_values.get(bundle_id, 0) or 0))
    return 0


def _style_header(row, fill):
    for cell in row:
        cell.fill = PatternFill("solid", fgColor=fill)
        cell.font = Font(name="Aptos", size=10, bold=True, color=_WHITE)
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        cell.border = Border(bottom=_BORDER)


def export_room_bundle_matrix(path, room_types, bundles, quantities) -> str:
    """Write an editable matrix plus normalized assignment/reference sheets."""
    destination = Path(path)
    if destination.suffix.casefold() != ".xlsx":
        destination = destination.with_suffix(".xlsx")
    rooms = [dict(row) for row in room_types if isinstance(row, dict)]
    bundle_rows = [dict(row) for row in bundles if isinstance(row, dict)]
    if not rooms:
        raise XlsxError("Select at least one room type to export.")
    if not bundle_rows:
        raise XlsxError("Select at least one asset bundle to export.")

    try:
        workbook = Workbook()
        matrix = workbook.active
        matrix.title = MATRIX_SHEET_NAME
        matrix.sheet_view.showGridLines = False
        last_column = 2 + len(bundle_rows)
        last_letter = get_column_letter(last_column)
        matrix.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_column)
        matrix["A1"] = "Room Type Asset Bundle Assignment Matrix"
        matrix["A1"].fill = PatternFill("solid", fgColor=_NAVY)
        matrix["A1"].font = Font(
            name="Aptos Display", size=15, bold=True, color=_WHITE
        )
        matrix["A1"].alignment = Alignment(vertical="center")
        matrix.row_dimensions[1].height = 28
        matrix.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_column)
        matrix["A2"] = (
            "Enter a positive whole-number quantity to apply a bundle; use 0 or "
            "blank for no assignment. Keep room and bundle IDs unchanged for import."
        )
        matrix["A2"].fill = PatternFill("solid", fgColor=_PALE_YELLOW)
        matrix["A2"].font = Font(name="Aptos", size=10, italic=True)
        matrix["A2"].alignment = Alignment(wrap_text=True, vertical="center")
        matrix.row_dimensions[2].height = 34

        headers = ["Room Type ID", "Room Type Name"] + [
            f"{_bundle_id(bundle, index)} | {_bundle_name(bundle)}"
            for index, bundle in enumerate(bundle_rows)
        ]
        matrix.append([])
        matrix.append(headers)
        _style_header(matrix[4], _BLUE)
        matrix.row_dimensions[4].height = 42
        for room_index, room in enumerate(rooms):
            room_id = _room_id(room, room_index)
            row = [room_id, _room_name(room)]
            for bundle_index, bundle in enumerate(bundle_rows):
                bundle_id = _bundle_id(bundle, bundle_index)
                value = _quantity_lookup(quantities, room_id, bundle_id)
                row.append(value if value > 0 else None)
            matrix.append(row)
        matrix.freeze_panes = "C5"
        matrix.auto_filter.ref = f"A4:{last_letter}{4 + len(rooms)}"
        matrix.column_dimensions["A"].width = 20
        matrix.column_dimensions["B"].width = 34
        for column in range(3, last_column + 1):
            matrix.column_dimensions[get_column_letter(column)].width = 22
        editable_range = f"C5:{last_letter}{4 + len(rooms)}"
        matrix[editable_range][0][0].number_format = "0"
        for row in matrix.iter_rows(
            min_row=5,
            max_row=4 + len(rooms),
            min_col=3,
            max_col=last_column,
        ):
            for cell in row:
                cell.fill = PatternFill("solid", fgColor=_PALE_BLUE)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.number_format = "0"
        validation = DataValidation(
            type="whole",
            operator="between",
            formula1="0",
            formula2="100000",
            allow_blank=True,
        )
        validation.error = "Enter a whole-number quantity from 0 to 100000."
        validation.errorTitle = "Invalid bundle quantity"
        validation.prompt = "Blank/0 removes the assignment; a positive value applies it."
        validation.promptTitle = "Bundle quantity"
        matrix.add_data_validation(validation)
        validation.add(editable_range)

        assignments = workbook.create_sheet(ROWS_SHEET_NAME)
        assignments.sheet_view.showGridLines = False
        assignments.append(
            ["Room Type ID", "Room Type Name", "Bundle ID", "Bundle Name", "Quantity"]
        )
        _style_header(assignments[1], _TEAL)
        for room_index, room in enumerate(rooms):
            room_id = _room_id(room, room_index)
            for bundle_index, bundle in enumerate(bundle_rows):
                bundle_id = _bundle_id(bundle, bundle_index)
                value = _quantity_lookup(quantities, room_id, bundle_id)
                if value <= 0:
                    continue
                assignments.append(
                    [
                        room_id,
                        _room_name(room),
                        bundle_id,
                        _bundle_name(bundle),
                        value,
                    ]
                )
        assignments.freeze_panes = "A2"
        assignments.auto_filter.ref = (
            f"A1:E{max(1, assignments.max_row)}"
        )
        for column, width in {"A": 20, "B": 34, "C": 20, "D": 34, "E": 12}.items():
            assignments.column_dimensions[column].width = width
        for cell in assignments["E"][1:]:
            cell.number_format = "0"

        reference = workbook.create_sheet(REFERENCE_SHEET_NAME)
        reference.sheet_view.showGridLines = False
        reference.append(["Bundle ID", "Bundle Name", "Description", "Asset lines"])
        _style_header(reference[1], _TEAL)
        for index, bundle in enumerate(bundle_rows):
            reference.append(
                [
                    _bundle_id(bundle, index),
                    _bundle_name(bundle),
                    _text(bundle.get("description")),
                    len(bundle.get("assets", []) or []),
                ]
            )
        reference.freeze_panes = "A2"
        reference.auto_filter.ref = f"A1:D{max(1, reference.max_row)}"
        for column, width in {"A": 20, "B": 34, "C": 55, "D": 14}.items():
            reference.column_dimensions[column].width = width
        for row in reference.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        workbook.properties.creator = "Cable Route Resolver"
        workbook.properties.lastModifiedBy = "Cable Route Resolver"
        workbook.save(destination)
    except (OSError, ValueError) as exc:
        raise XlsxError(f"Could not write {destination.name}: {exc}") from exc
    return str(destination)


def _aliases(rows, id_getter, name_getter):
    by_alias = {}
    for index, row in enumerate(rows):
        identity = id_getter(row, index)
        name = name_getter(row)
        aliases = {
            _key(identity),
            _key(name),
            _key(f"{identity} | {name}"),
            _key(f"{identity} - {name}"),
            _key(f"{name} [{identity}]"),
        }
        for alias in aliases:
            if alias:
                by_alias.setdefault(alias, identity)
    return by_alias


def _match_header(value, aliases):
    key = _key(value)
    if key in aliases:
        return aliases[key]
    text = _text(value)
    for pattern in (r"^\s*([^|\[]+)\s*\|", r"\[([^\]]+)\]\s*$"):
        match = re.search(pattern, text)
        if match and _key(match.group(1)) in aliases:
            return aliases[_key(match.group(1))]
    return ""


def _header_index(headers, candidates):
    candidate_keys = {_key(value) for value in candidates}
    for index, value in enumerate(headers):
        if _key(value) in candidate_keys:
            return index
    return -1


def import_room_bundle_matrix(path, room_types, bundles) -> dict:
    """Read canonical or similar matrix/long-form assignment workbooks."""
    rooms = [dict(row) for row in room_types if isinstance(row, dict)]
    bundle_rows = [dict(row) for row in bundles if isinstance(row, dict)]
    room_aliases = _aliases(rooms, _room_id, _room_name)
    bundle_aliases = _aliases(bundle_rows, _bundle_id, _bundle_name)
    workbook = read_xlsx(path)
    warnings = []
    canonical_matrix_available = False
    for sheet in workbook.sheets:
        if sheet.name.casefold() != MATRIX_SHEET_NAME.casefold():
            continue
        for row in sheet.rows[:30]:
            has_room = max(
                _header_index(
                    row, {"room type id", "room id", "room type code"}
                ),
                _header_index(
                    row, {"room type name", "room type", "room"}
                ),
            ) >= 0
            has_bundle = any(
                _match_header(value, bundle_aliases) for value in row
            )
            if has_room and has_bundle:
                canonical_matrix_available = True
                break

    # Prefer a recognizable long-form assignment table when present.
    for sheet in workbook.sheets:
        if (
            canonical_matrix_available
            and sheet.name.casefold() == ROWS_SHEET_NAME.casefold()
        ):
            continue
        for header_index, row in enumerate(sheet.rows[:20]):
            room_id_col = _header_index(
                row, {"room type id", "room id", "room type code"}
            )
            room_name_col = _header_index(row, {"room type name", "room type", "room"})
            bundle_id_col = _header_index(
                row,
                {
                    "bundle id",
                    "asset bundle id",
                    "use case id",
                    "use-case id",
                    "uc id",
                },
            )
            bundle_name_col = _header_index(
                row,
                {
                    "bundle name",
                    "asset bundle",
                    "bundle",
                    "use case",
                    "use case name",
                },
            )
            quantity_col = _header_index(
                row,
                {
                    "quantity",
                    "qty",
                    "bundle qty",
                    "bundle quantity",
                    "count",
                    "instances",
                    "applies",
                },
            )
            if (
                max(room_id_col, room_name_col) < 0
                or max(bundle_id_col, bundle_name_col) < 0
                or quantity_col < 0
            ):
                continue
            assignments = defaultdict(dict)
            matched_rooms = set()
            matched_bundles = set()
            for excel_row, values in enumerate(
                sheet.rows[header_index + 1 :], start=header_index + 2
            ):
                room_value = (
                    values[room_id_col]
                    if 0 <= room_id_col < len(values)
                    else ""
                )
                room_id = _match_header(room_value, room_aliases)
                if not room_id and 0 <= room_name_col < len(values):
                    room_id = _match_header(values[room_name_col], room_aliases)
                bundle_value = (
                    values[bundle_id_col]
                    if 0 <= bundle_id_col < len(values)
                    else ""
                )
                bundle_id = _match_header(bundle_value, bundle_aliases)
                if not bundle_id and 0 <= bundle_name_col < len(values):
                    bundle_id = _match_header(
                        values[bundle_name_col], bundle_aliases
                    )
                if not room_id or not bundle_id:
                    continue
                value = (
                    values[quantity_col]
                    if quantity_col < len(values)
                    else ""
                )
                quantity = _quantity(value)
                if quantity is None:
                    warnings.append(
                        f"{sheet.name} row {excel_row}: ignored invalid quantity "
                        f"{_text(value)!r}."
                    )
                    continue
                matched_rooms.add(room_id)
                matched_bundles.add(bundle_id)
                assignments[room_id][bundle_id] = (
                    assignments[room_id].get(bundle_id, 0) + quantity
                )
            if matched_rooms and matched_bundles:
                return {
                    "assignments": {
                        room_id: dict(values)
                        for room_id, values in assignments.items()
                    },
                    "matched_room_ids": sorted(matched_rooms),
                    "matched_bundle_ids": sorted(matched_bundles),
                    "warnings": warnings,
                    "sheet": sheet.name,
                    "layout": "rows",
                }

    # Otherwise locate a matrix header row containing room and bundle labels.
    for sheet in workbook.sheets:
        for header_index, row in enumerate(sheet.rows[:30]):
            room_id_col = _header_index(
                row, {"room type id", "room id", "room type code"}
            )
            room_name_col = _header_index(row, {"room type name", "room type", "room"})
            if max(room_id_col, room_name_col) < 0:
                continue
            bundle_columns = {
                column: bundle_id
                for column, value in enumerate(row)
                if (
                    bundle_id := _match_header(value, bundle_aliases)
                )
            }
            if not bundle_columns:
                continue
            assignments = defaultdict(dict)
            matched_rooms = set()
            for excel_row, values in enumerate(
                sheet.rows[header_index + 1 :], start=header_index + 2
            ):
                room_id = ""
                if 0 <= room_id_col < len(values):
                    room_id = _match_header(values[room_id_col], room_aliases)
                if not room_id and 0 <= room_name_col < len(values):
                    room_id = _match_header(values[room_name_col], room_aliases)
                if not room_id:
                    continue
                matched_rooms.add(room_id)
                for column, bundle_id in bundle_columns.items():
                    value = values[column] if column < len(values) else ""
                    quantity = _quantity(value)
                    if quantity is None:
                        warnings.append(
                            f"{sheet.name} row {excel_row}, column "
                            f"{get_column_letter(column + 1)}: ignored invalid "
                            f"quantity {_text(value)!r}."
                        )
                        continue
                    assignments[room_id][bundle_id] = quantity
            if matched_rooms:
                return {
                    "assignments": {
                        room_id: dict(values)
                        for room_id, values in assignments.items()
                    },
                    "matched_room_ids": sorted(matched_rooms),
                    "matched_bundle_ids": sorted(set(bundle_columns.values())),
                    "warnings": warnings,
                    "sheet": sheet.name,
                    "layout": "matrix",
                }

    # Also accept transposed matrices with bundles/use cases down the rows.
    for sheet in workbook.sheets:
        for header_index, row in enumerate(sheet.rows[:30]):
            bundle_id_col = _header_index(
                row,
                {
                    "bundle id",
                    "asset bundle id",
                    "use case id",
                    "use-case id",
                    "uc id",
                },
            )
            bundle_name_col = _header_index(
                row,
                {
                    "bundle name",
                    "asset bundle",
                    "bundle",
                    "use case",
                    "use case name",
                },
            )
            if max(bundle_id_col, bundle_name_col) < 0:
                continue
            room_columns = {
                column: room_id
                for column, value in enumerate(row)
                if (room_id := _match_header(value, room_aliases))
            }
            if not room_columns:
                continue
            assignments = defaultdict(dict)
            matched_bundles = set()
            for excel_row, values in enumerate(
                sheet.rows[header_index + 1 :], start=header_index + 2
            ):
                bundle_id = ""
                if 0 <= bundle_id_col < len(values):
                    bundle_id = _match_header(
                        values[bundle_id_col], bundle_aliases
                    )
                if not bundle_id and 0 <= bundle_name_col < len(values):
                    bundle_id = _match_header(
                        values[bundle_name_col], bundle_aliases
                    )
                if not bundle_id:
                    continue
                matched_bundles.add(bundle_id)
                for column, room_id in room_columns.items():
                    value = values[column] if column < len(values) else ""
                    quantity = _quantity(value)
                    if quantity is None:
                        warnings.append(
                            f"{sheet.name} row {excel_row}, column "
                            f"{get_column_letter(column + 1)}: ignored invalid "
                            f"quantity {_text(value)!r}."
                        )
                        continue
                    assignments[room_id][bundle_id] = quantity
            if matched_bundles:
                return {
                    "assignments": {
                        room_id: dict(values)
                        for room_id, values in assignments.items()
                    },
                    "matched_room_ids": sorted(set(room_columns.values())),
                    "matched_bundle_ids": sorted(matched_bundles),
                    "warnings": warnings,
                    "sheet": sheet.name,
                    "layout": "transposed matrix",
                }

    raise XlsxError(
        "No room-type/bundle assignment matrix was recognized. Include a room "
        "type ID or name column and bundle IDs/names as column headers, or use "
        "long-form Room Type, Bundle, and Quantity columns."
    )
