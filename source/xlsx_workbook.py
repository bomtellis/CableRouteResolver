"""Excel workbook helpers backed by openpyxl."""

from __future__ import annotations

from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
import re
from typing import Sequence

from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.exceptions import InvalidFileException
from zipfile import BadZipFile


class XlsxError(ValueError):
    """Raised when an Excel workbook cannot be read or written."""


@dataclass(frozen=True)
class SheetData:
    name: str
    rows: tuple[tuple[object, ...], ...]

    @property
    def max_columns(self) -> int:
        return max((len(row) for row in self.rows), default=0)


@dataclass(frozen=True)
class WorkbookData:
    path: str
    sheets: tuple[SheetData, ...]

    @property
    def sheet_names(self) -> tuple[str, ...]:
        return tuple(sheet.name for sheet in self.sheets)

    def sheet(self, name: str) -> SheetData:
        for sheet in self.sheets:
            if sheet.name == name:
                return sheet
        raise XlsxError(f'Worksheet "{name}" was not found.')


@dataclass
class WorksheetSpec:
    name: str
    rows: Sequence[Sequence[object]]
    title_rows: set[int] = field(default_factory=set)
    section_rows: set[int] = field(default_factory=set)
    header_rows: set[int] = field(default_factory=set)
    status_column: int | None = None
    freeze_row: int = 0
    auto_filter_row: int | None = None


def _trim_row(values: Sequence[object]) -> tuple[object, ...]:
    result = list(values)
    while result and result[-1] in (None, ""):
        result.pop()
    return tuple("" if value is None else value for value in result)


def read_xlsx(path: str | Path) -> WorkbookData:
    """Read cached worksheet values from an XLSX or XLSM workbook."""

    source = Path(path)
    if source.suffix.casefold() not in {".xlsx", ".xlsm"}:
        raise XlsxError("Only .xlsx and .xlsm workbooks are supported.")
    try:
        cached = load_workbook(
            source, read_only=True, data_only=True, keep_links=False
        )
        formulas = load_workbook(
            source, read_only=True, data_only=False, keep_links=False
        )
        sheets = []
        for cached_sheet in cached.worksheets:
            formula_sheet = formulas[cached_sheet.title]
            rows = []
            for cached_row, formula_row in zip(
                cached_sheet.iter_rows(values_only=True),
                formula_sheet.iter_rows(values_only=True),
            ):
                values = []
                for cached_value, formula_value in zip(cached_row, formula_row):
                    value = cached_value
                    if value is None and isinstance(formula_value, str):
                        value = formula_value
                    values.append(value)
                rows.append(_trim_row(values))
            while rows and not rows[-1]:
                rows.pop()
            sheets.append(SheetData(cached_sheet.title, tuple(rows)))
        cached.close()
        formulas.close()
    except (OSError, ValueError, KeyError, BadZipFile, InvalidFileException) as exc:
        raise XlsxError(f"Could not read {source.name}: {exc}") from exc
    if not sheets:
        raise XlsxError(f"{source.name} does not contain any readable worksheets.")
    return WorkbookData(str(source), tuple(sheets))


_NAVY = "17365D"
_TEAL = "0F6B78"
_BLUE = "24557A"
_GREEN = "D1E7DD"
_RED = "F8D7DA"
_AMBER = "FFF3CD"
_GRAY = "E9ECEF"
_WHITE = "FFFFFF"
_BORDER = Side(style="thin", color="D9E2F3")


def _sheet_name(name: str, used: set[str], fallback: str) -> str:
    base = re.sub(r"[\[\]:*?/\\]", "_", name).strip("' ")[:31] or fallback
    candidate = base
    suffix = 2
    while candidate.casefold() in used:
        tail = f" ({suffix})"
        candidate = f"{base[:31 - len(tail)]}{tail}"
        suffix += 1
    used.add(candidate.casefold())
    return candidate


def _excel_value(value: object) -> object:
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, (date, datetime, int, float, bool)) or value is None:
        return value
    text = str(value)
    text = ILLEGAL_CHARACTERS_RE.sub("", text)
    # Comparison reports should show unresolved source formulas as text rather
    # than creating broken cross-sheet formulas in the new workbook.
    return f"'{text}" if text.startswith("=") else text


def _style_row(worksheet, row_number: int, spec: WorksheetSpec) -> None:
    index = row_number - 1
    cells = worksheet[row_number]
    fill = None
    font = Font(name="Aptos", size=10, color="212529")
    if index in spec.title_rows:
        fill = PatternFill("solid", fgColor=_NAVY)
        font = Font(name="Aptos Display", size=14, bold=True, color=_WHITE)
    elif index in spec.section_rows:
        fill = PatternFill("solid", fgColor=_TEAL)
        font = Font(name="Aptos", size=10, bold=True, color=_WHITE)
    elif index in spec.header_rows:
        fill = PatternFill("solid", fgColor=_BLUE)
        font = Font(name="Aptos", size=10, bold=True, color=_WHITE)
    elif spec.status_column is not None:
        status_cell = worksheet.cell(row_number, spec.status_column + 1)
        status = str(status_cell.value or "").casefold()
        color = {
            "added": _GREEN,
            "removed": _RED,
            "quantity changed": _AMBER,
            "changed": _AMBER,
            "same": _GRAY,
        }.get(status)
        if color:
            fill = PatternFill("solid", fgColor=color)
    for cell in cells:
        cell.font = font
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        if isinstance(cell.value, (int, float, Decimal)) and not isinstance(
            cell.value, bool
        ):
            cell.number_format = "#,##0.###"
        elif isinstance(cell.value, (date, datetime)):
            cell.number_format = "yyyy-mm-dd"
        if fill:
            cell.fill = fill
        if index in spec.header_rows:
            cell.border = Border(bottom=_BORDER)
    if index in spec.title_rows | spec.section_rows | spec.header_rows:
        worksheet.row_dimensions[row_number].height = 24


def write_xlsx(path: str | Path, worksheets: Sequence[WorksheetSpec]) -> str:
    """Write a styled XLSX report and return its final path."""

    if not worksheets:
        raise XlsxError("At least one worksheet is required.")
    destination = Path(path)
    if destination.suffix.casefold() != ".xlsx":
        destination = destination.with_suffix(".xlsx")
    try:
        destination.parent.mkdir(parents=True, exist_ok=True)
        workbook = Workbook()
        workbook.remove(workbook.active)
        used_names: set[str] = set()
        for index, spec in enumerate(worksheets, start=1):
            worksheet = workbook.create_sheet(
                _sheet_name(spec.name, used_names, f"Sheet {index}")
            )
            worksheet.sheet_view.showGridLines = False
            for row in spec.rows:
                worksheet.append([_excel_value(value) for value in row])
            for row_number in range(1, worksheet.max_row + 1):
                _style_row(worksheet, row_number, spec)
            for column in range(1, worksheet.max_column + 1):
                width = max(
                    (
                        len(str(worksheet.cell(row, column).value or ""))
                        for row in range(1, min(worksheet.max_row, 2000) + 1)
                    ),
                    default=0,
                )
                worksheet.column_dimensions[get_column_letter(column)].width = min(
                    55, max(10, width + 2)
                )
            if spec.freeze_row > 0:
                worksheet.freeze_panes = f"A{spec.freeze_row + 1}"
            if spec.auto_filter_row is not None and worksheet.max_row:
                first = spec.auto_filter_row + 1
                worksheet.auto_filter.ref = (
                    f"A{first}:{get_column_letter(worksheet.max_column)}"
                    f"{worksheet.max_row}"
                )
        workbook.properties.creator = "Cable Route Resolver"
        workbook.properties.lastModifiedBy = "Cable Route Resolver"
        workbook.save(destination)
    except (OSError, ValueError) as exc:
        raise XlsxError(f"Could not write {destination.name}: {exc}") from exc
    return str(destination)
