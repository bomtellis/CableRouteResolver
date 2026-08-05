"""Regression tests for the Excel workbook comparison workflow."""

import sys
import tempfile
import unittest
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook


sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "source"))

from workbook_compare import (
    RoomMappingGroup,
    SchemaMapping,
    compare_workbooks,
    expand_assets,
    export_comparison,
    extract_rooms,
    suggest_schema,
)
from xlsx_workbook import SheetData, WorkbookData


MAIN_HEADERS = (
    "Room Code",
    "Room Name",
    "Code",
    "Description",
    "Count",
    "Assembly ID",
)
ASSEMBLY_HEADERS = (
    "Assembly ID",
    "Parent Code",
    "Item Quantity",
    "Item Code",
    "Item Description",
)


def workbook(path, main_rows, assembly_rows=()):
    sheets = [
        SheetData("Equipment List", (MAIN_HEADERS, *tuple(main_rows))),
    ]
    if assembly_rows:
        sheets.append(
            SheetData("Assembly Items", (ASSEMBLY_HEADERS, *tuple(assembly_rows)))
        )
    return WorkbookData(path, tuple(sheets))


def schema(with_assemblies=True):
    return SchemaMapping(
        main_sheet="Equipment List",
        main_header_row=0,
        main_columns={
            "room_code": 0,
            "room_name": 1,
            "asset_code": 2,
            "description": 3,
            "quantity": 4,
            "assembly_id": 5,
            "assembly_parent_code": None,
        },
        assembly_sheet="Assembly Items" if with_assemblies else None,
        assembly_header_row=0,
        assembly_columns=(
            {
                "assembly_id": 0,
                "parent_code": 1,
                "item_quantity": 2,
                "item_code": 3,
                "item_description": 4,
                "room_code": None,
                "room_name": None,
            }
            if with_assemblies
            else {}
        ),
    )


class WorkbookComparisonTests(unittest.TestCase):
    def test_schema_detection_prefers_equipment_and_assembly_sheets(self):
        source = workbook(
            "source.xlsx",
            [("R1", "Room one", "A-ASM", "Assembly", 1, "ASM-1")],
            [("ASM-1", "A-ASM", 1, "ITEM-1", "Child")],
        )
        detected = suggest_schema(source)
        self.assertEqual(detected.main_sheet, "Equipment List")
        self.assertEqual(detected.assembly_sheet, "Assembly Items")
        self.assertEqual(detected.main_columns["room_name"], 1)
        self.assertEqual(detected.assembly_columns["item_code"], 3)

    def test_schema_detection_maps_alternate_column_names_without_assembly(self):
        source = WorkbookData(
            "alternate.xlsx",
            (
                SheetData(
                    "Schedule v2",
                    (
                        ("Space Name", "Room Type Code", "Equipment Code", "Qty", "Equipment Description"),
                        ("Office", "OF-1", "CHAIR", 1, "Chair"),
                    ),
                ),
            ),
        )
        detected = suggest_schema(source)
        self.assertEqual(detected.main_sheet, "Schedule v2")
        self.assertIsNone(detected.assembly_sheet)
        self.assertEqual(detected.main_columns["room_name"], 0)
        self.assertEqual(detected.main_columns["room_code"], 1)
        self.assertEqual(detected.main_columns["asset_code"], 2)
        self.assertEqual(detected.main_columns["quantity"], 3)

    def test_assembly_items_are_expanded_and_parent_quantity_multiplied(self):
        source = workbook(
            "source.xlsx",
            [
                ("R1", "Room one", "A-ASM", "Assembly", 2, "ASM-1"),
                ("R1", "Room one", "CHAIR", "Chair", 3, ""),
            ],
            [
                ("ASM-1", "A-ASM", 4, "OUTLET", "Twin outlet"),
                ("ASM-1", "A-ASM", 1, "PANEL", "Panel"),
            ],
        )
        room_key = next(iter(extract_rooms(source, schema())))
        assets, warnings = expand_assets(source, schema(), [room_key])
        quantities = {asset.asset_code: asset.quantity for asset in assets}
        self.assertEqual(quantities["OUTLET"], Decimal(8))
        self.assertEqual(quantities["PANEL"], Decimal(2))
        self.assertEqual(quantities["CHAIR"], Decimal(3))
        assembly_rows = [asset for asset in assets if asset.assembly_id == "ASM-1"]
        self.assertEqual({asset.assembly_id for asset in assembly_rows}, {"ASM-1"})
        self.assertFalse(warnings)

    def test_no_assembly_sheet_keeps_parent_as_direct_asset(self):
        source = workbook(
            "source.xlsx",
            [("R1", "Room one", "A-ASM", "Assembly", 2, "ASM-1")],
        )
        room_key = next(iter(extract_rooms(source, schema(False))))
        assets, warnings = expand_assets(source, schema(False), [room_key])
        self.assertEqual(len(assets), 1)
        self.assertEqual(assets[0].asset_code, "A-ASM")
        self.assertEqual(assets[0].quantity, Decimal(2))
        self.assertEqual(assets[0].assembly_id, "")
        self.assertFalse(warnings)

    def test_many_to_one_room_mapping_aggregates_quantities(self):
        source_a = workbook(
            "a.xlsx",
            [("A1", "Single room", "CHAIR", "Chair", 3, "")],
        )
        source_b = workbook(
            "b.xlsx",
            [
                ("B1", "First room", "CHAIR", "Chair", 1, ""),
                ("B2", "Second room", "CHAIR", "Chair", 3, ""),
            ],
        )
        rooms_a = extract_rooms(source_a, schema(False))
        rooms_b = extract_rooms(source_b, schema(False))
        result = compare_workbooks(
            source_a,
            schema(False),
            source_b,
            schema(False),
            [
                RoomMappingGroup(
                    tuple(rooms_a),
                    tuple(rooms_b),
                    "Combined comparison",
                )
            ],
            include_unmapped=False,
        )
        row = result.mappings[0].total_rows[0]
        self.assertEqual(row.quantity_a, Decimal(3))
        self.assertEqual(row.quantity_b, Decimal(4))
        self.assertEqual(row.difference, Decimal(1))
        self.assertEqual(row.status, "Quantity changed")

    def test_export_contains_both_originals_and_differences_tabs(self):
        source_a = workbook(
            "a.xlsx",
            [("A1", "Room A", "CHAIR", "Chair", 1, "")],
        )
        source_b = workbook(
            "b.xlsx",
            [("B1", "Room B", "CHAIR", "Chair", 2, "")],
        )
        rooms_a = extract_rooms(source_a, schema(False))
        rooms_b = extract_rooms(source_b, schema(False))
        result = compare_workbooks(
            source_a,
            schema(False),
            source_b,
            schema(False),
            [RoomMappingGroup(tuple(rooms_a), tuple(rooms_b), "Room comparison")],
            include_unmapped=False,
        )
        with tempfile.TemporaryDirectory() as directory:
            destination = Path(directory) / "comparison.xlsx"
            export_comparison(
                destination,
                result,
                source_a,
                schema(False),
                source_b,
                schema(False),
            )
            exported = load_workbook(destination, data_only=True)
            self.assertEqual(
                exported.sheetnames, ["Original A", "Original B", "Differences"]
            )
            differences = exported["Differences"]
            self.assertEqual(differences["N6"].value, "Quantity changed")
            self.assertEqual(differences.freeze_panes, "A6")
            exported.close()


if __name__ == "__main__":
    unittest.main()
